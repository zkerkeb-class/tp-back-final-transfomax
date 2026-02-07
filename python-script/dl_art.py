import openpyxl
import os
import re
import requests

def dl_image(url, number):

    path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    path = os.path.join(path, "assets", "pokemons", "full-art")

    try:
        response = requests.get(url, stream=True)
        
        if response.status_code == 200:
            with open(os.path.join(path, f"{number}.png"), 'wb') as f:
                for chunk in response.iter_content(1024):
                    f.write(chunk)
            print(f"Succès : Image enregistrée sous {path}")
        else:
            print(f"Erreur : Impossible de télécharger (Code {response.status_code})")
            
    except Exception as e:
        print(f"Erreur lors du téléchargement : {e}")


dossier_script = os.path.dirname(os.path.abspath(__file__))
fichier_excel = os.path.join(dossier_script, "151_art.xlsx")

print(f"Analyse du fichier : {fichier_excel}...")

try:
    wb = openpyxl.load_workbook(fichier_excel)
    ws = wb.active 


    for row in ws.iter_rows(min_row=2):
        id, image = ws.cell(row=row[0].row, column=3).value, ws.cell(row=row[0].row, column=13).value

        if image:
            image = re.search(r'image\("(.+?)"\)', image)

            if image:
                dl_image(image.group(1), int(id))

            else :
                continue

        else :
            continue

except Exception as e:
    print(f"Erreur : {e}")